"""Esportazione per il commercialista: registro, riepilogo IVA, copie PDF.

Il difetto che questi test presidiano: gli importi finivano nell'Excel come
**testo**. Il file si apriva, sembrava giusto, e selezionando la colonna Totale
non compariva nessuna somma — cioe' veniva a mancare l'unico motivo per dare un
.xlsx invece del CSV. Un difetto che non da' errore e che si scopre solo quando
il commercialista ci lavora sopra.
"""
import io
import zipfile
from datetime import date
from decimal import Decimal

import openpyxl
import pytest

from app import export_commercialista as exc
from app.calcolo import RigaInput
from app.db import get_conn, init_db
from app.fatturazione import emetti_fattura

DAL, AL = "2026-01-01", "2026-12-31"


@pytest.fixture()
def conn(tmp_path):
    """Un periodo realistico: due fatture e una nota di credito che ne storna una."""
    db = tmp_path / "test.db"
    init_db(db)
    c = get_conn(db)
    c.execute(
        "INSERT INTO clienti (tipo, nome, cognome, codice_fiscale) "
        "VALUES ('fisica','Mario','Rossi','RSSMRA85T10A562S')"
    )
    c.commit()
    cliente = c.execute("SELECT * FROM clienti WHERE id=1").fetchone()

    prima = emetti_fattura(
        c, cliente=cliente, data_emissione="2026-02-15",
        righe=[RigaInput("Visita clinica equino", 1, "80", "22")],
        stato="incassata", data_pagamento="2026-02-15",
    )
    emetti_fattura(
        c, cliente=cliente, data_emissione="2026-03-10",
        righe=[RigaInput("Radiografia", 1, "120", "22")],
    )
    emetti_fattura(
        c, cliente=cliente, data_emissione="2026-04-01", tipo_documento="nota_credito",
        righe=[RigaInput("Storno visita", 1, "80", "22")],
        documento_riferimento_id=prima["id"],
    )
    yield c
    c.close()


def _foglio(conn, nome: str):
    wb = openpyxl.load_workbook(io.BytesIO(exc.genera_xlsx(conn, DAL, AL)))
    return wb[nome]


# --- il difetto vero -------------------------------------------------------

def test_gli_importi_dell_excel_sono_numeri_sommabili(conn):
    """**Il requisito.** Selezionando una colonna di importi Excel deve dare la somma.

    Con celle di testo non somma niente, e il commercialista se ne accorge solo
    quando ci sta gia' lavorando.
    """
    ws = _foglio(conn, "Registro")
    intestazioni = [c.value for c in ws[1]]
    colonne_importo = ["Imponibile", "Contributo ENPAV", "IVA", "Totale", "Ritenuta"]

    for nome in colonne_importo:
        i = intestazioni.index(nome) + 1
        for riga in range(2, ws.max_row + 1):
            cella = ws.cell(riga, i)
            assert cella.data_type == "n", \
                f"{nome} riga {riga}: cella di testo {cella.value!r}, Excel non la somma"
            assert isinstance(cella.value, (int, float))


def test_le_date_dell_excel_sono_date_vere(conn):
    """Come testo non si ordinano ne' si filtrano per mese."""
    ws = _foglio(conn, "Registro")
    i = [c.value for c in ws[1]].index("Data") + 1
    for riga in range(2, ws.max_row + 1):
        cella = ws.cell(riga, i)
        assert cella.is_date, f"riga {riga}: data come testo {cella.value!r}"
    assert ws.cell(2, i).value.date() == date(2026, 2, 15)


def test_anche_il_riepilogo_iva_e_numerico(conn):
    """I due fogli devono comportarsi allo stesso modo: prima non era cosi'."""
    ws = _foglio(conn, "Riepilogo IVA")
    for riga in range(2, ws.max_row + 1):
        for col in range(2, 6):   # tutto tranne l'aliquota
            assert ws.cell(riga, col).data_type == "n"


def test_l_intestazione_resta_visibile_scorrendo(conn):
    """Un registro annuale e' lungo: senza blocco si perde di vista cosa si legge."""
    assert _foglio(conn, "Registro").freeze_panes == "A2"


# --- il CSV non deve cambiare ----------------------------------------------

def test_il_csv_resta_identico_nel_formato(conn):
    """La tipizzazione e' per l'Excel: il CSV deve restare quello che era.

    Il commercialista potrebbe avere una procedura che lo legge.
    """
    testo = exc.genera_csv(conn, DAL, AL).decode("utf-8-sig")
    righe = testo.splitlines()
    assert exc.genera_csv(conn, DAL, AL)[:3] == b"\xef\xbb\xbf"   # BOM per Excel italiano
    assert righe[0] == ";".join(exc.COLONNE)
    # importi con due decimali e data ISO, come prima della tipizzazione
    assert righe[1].startswith("1/2026;2026-02-15;Fattura;Rossi Mario;RSSMRA85T10A562S;;80.00;")
    assert "Riepilogo IVA per aliquota" in testo


# --- contenuto contabile ---------------------------------------------------

def test_la_nota_di_credito_ha_segno_negativo(conn):
    righe = exc.registro(conn, DAL, AL)
    nota = next(r for r in righe if r[2] == "Nota credito")
    assert nota[6] < 0 and nota[10] < 0, "la nota di credito deve sottrarre"


def test_il_riepilogo_iva_quadra_col_registro(conn):
    """Se non tornano, uno dei due sta mentendo al commercialista."""
    righe = exc.registro(conn, DAL, AL)
    somma_registro = sum(r[9] for r in righe)                    # colonna IVA
    somma_riepilogo = sum(g["iva"] for g in exc.riepilogo_iva(conn, DAL, AL))
    assert somma_registro == somma_riepilogo

    imponibile_registro = sum(r[6] for r in righe)
    imponibile_riepilogo = sum(g["imponibile"] for g in exc.riepilogo_iva(conn, DAL, AL))
    assert imponibile_registro == imponibile_riepilogo
    # 80 + 120 - 80 storno = 120
    assert imponibile_registro == Decimal("120.00")


def test_il_registro_restituisce_valori_tipizzati(conn):
    """Chi scrive il file decide come renderli: il registro non pre-formatta."""
    prima = exc.registro(conn, DAL, AL)[0]
    assert isinstance(prima[1], date)
    assert all(isinstance(prima[i], Decimal) for i in (6, 7, 9, 10, 11))


def test_lo_zip_contiene_un_pdf_valido_per_documento(conn):
    """Note di credito comprese: senza, la contabilita' del periodo non torna."""
    z = zipfile.ZipFile(io.BytesIO(exc.genera_zip_pdf(conn, DAL, AL, {})))
    nomi = z.namelist()
    assert len(nomi) == 3
    assert any(n.startswith("nota_credito_") for n in nomi)
    for n in nomi:
        assert z.read(n)[:4] == b"%PDF", f"{n} non e' un PDF"


def test_fuori_periodo_non_entra_nulla(conn):
    assert exc.registro(conn, "2025-01-01", "2025-12-31") == []
