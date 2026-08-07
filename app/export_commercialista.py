"""Esportazioni per il commercialista.

- Registro delle fatture emesse nel periodo (per DATA DI EMISSIONE), in CSV e in
  Excel, con tutte le colonne fiscali e un riepilogo IVA per aliquota.
- Bundle ZIP con le copie PDF delle fatture del periodo.

Le note di credito compaiono con importi di segno negativo, cosi' il riepilogo
IVA rappresenta l'imposta netta del periodo.

**I valori restano tipizzati.** ``registro()`` produce ``Decimal`` per gli importi
e ``date`` per le date, non stringhe gia' formattate: la resa spetta a chi scrive
il file. Nel CSV ``csv.writer`` li rende identici a prima (``str(Decimal('95.00'))``
e' ``'95.00'``); nell'Excel diventano celle numeriche e celle data, che si possono
sommare e ordinare. Prima erano testo in entrambi, e il commercialista che
selezionava la colonna Totale non otteneva nessuna somma — cioe' mancava l'unico
motivo per consegnare un .xlsx invece del CSV.
"""
from __future__ import annotations

import csv
import io
import zipfile
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font

from app.fatturazione import gruppi_iva_da_righe, leggi_fattura
from app.pdf_fattura import genera_pdf_fattura

COLONNE = [
    "Numero", "Data", "Tipo", "Cliente", "CF", "P.IVA", "Imponibile",
    "Contributo ENPAV", "Aliquote IVA", "IVA", "Totale", "Ritenuta", "Stato",
]


def _segno(tipo: str) -> Decimal:
    return Decimal("-1") if tipo == "nota_credito" else Decimal("1")


def _data(valore: str) -> date | str:
    """Data ISO del database come oggetto ``date``.

    Se il valore non e' una data valida viene restituito com'e': meglio una cella
    con dentro il testo originale che un export che si rifiuta di generarsi.
    """
    try:
        return datetime.strptime((valore or "").strip(), "%Y-%m-%d").date()
    except ValueError:
        return valore


def _documenti(conn, dal: str, al: str) -> list[dict]:
    righe = conn.execute(
        "SELECT * FROM fatture "
        "WHERE stato != 'annullata' AND data_emissione BETWEEN ? AND ? "
        "ORDER BY data_emissione, tipo_documento, numero_progressivo",
        (dal, al),
    ).fetchall()
    return [dict(r) for r in righe]


def registro(conn, dal: str, al: str) -> list[list]:
    """Righe del registro (liste allineate a COLONNE), con segno per le note credito.

    Importi come ``Decimal`` e date come ``date``: vedi la nota in testa al modulo.
    """
    out: list[list] = []
    for f in _documenti(conn, dal, al):
        f["righe"] = leggi_fattura(conn, f["id"])["righe"]
        gruppi = gruppi_iva_da_righe(f["righe"], f["enpav_pct"])
        aliquote = ";".join(sorted({g["aliquota"].rstrip("0").rstrip(".") for g in gruppi}))
        s = _segno(f["tipo_documento"])
        out.append([
            f["numero_visualizzato"],
            _data(f["data_emissione"]),
            "Nota credito" if f["tipo_documento"] == "nota_credito" else "Fattura",
            f["cli_denominazione"],
            f["cli_codice_fiscale"],
            f["cli_partita_iva"],
            s * Decimal(f["imponibile"]),
            s * Decimal(f["contributo_enpav"]),
            aliquote,
            s * Decimal(f["iva_totale"]),
            s * Decimal(f["totale_documento"]),
            s * Decimal(f["ritenuta_importo"]),
            f["stato"],
        ])
    return out


def riepilogo_iva(conn, dal: str, al: str) -> list[dict]:
    """Aggrega imponibile/ENPAV/base/IVA per aliquota nel periodo (segno incluso)."""
    acc: dict[str, dict[str, Decimal]] = {}
    for f in _documenti(conn, dal, al):
        righe = leggi_fattura(conn, f["id"])["righe"]
        s = _segno(f["tipo_documento"])
        for g in gruppi_iva_da_righe(righe, f["enpav_pct"]):
            k = g["aliquota"].rstrip("0").rstrip(".")
            a = acc.setdefault(k, {"imponibile": Decimal("0"), "enpav": Decimal("0"),
                                   "base_iva": Decimal("0"), "iva": Decimal("0")})
            a["imponibile"] += s * Decimal(g["imponibile"])
            a["enpav"] += s * Decimal(g["enpav"])
            a["base_iva"] += s * Decimal(g["base_iva"])
            a["iva"] += s * Decimal(g["iva"])
    return [
        {"aliquota": k, "imponibile": v["imponibile"], "enpav": v["enpav"],
         "base_iva": v["base_iva"], "iva": v["iva"]}
        for k, v in sorted(acc.items())
    ]


def genera_csv(conn, dal: str, al: str) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", lineterminator="\r\n")
    w.writerow(COLONNE)
    for r in registro(conn, dal, al):
        w.writerow(r)
    w.writerow([])
    w.writerow(["Riepilogo IVA per aliquota"])
    w.writerow(["Aliquota %", "Imponibile", "ENPAV", "Base IVA", "IVA"])
    for g in riepilogo_iva(conn, dal, al):
        w.writerow([g["aliquota"], g["imponibile"], g["enpav"], g["base_iva"], g["iva"]])
    # BOM utf-8 per apertura corretta in Excel italiano.
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


# Colonne del registro che contengono importi (0-based su COLONNE).
_COLONNE_IMPORTO = {6, 7, 9, 10, 11}
_COLONNA_DATA = 1

_FORMATO_IMPORTO = "#,##0.00"
_FORMATO_DATA = "DD/MM/YYYY"


def _scrivi_riga(ws, valori: list) -> None:
    """Appende una riga convertendo i Decimal in numeri e applicando i formati.

    openpyxl scriverebbe un ``Decimal`` come testo: va passato a ``float``. La
    perdita di precisione non e' un problema qui — sono importi a due decimali,
    ben dentro l'esatto rappresentabile — mentre una cella di testo lo sarebbe,
    perche' non si somma.
    """
    ws.append([float(v) if isinstance(v, Decimal) else v for v in valori])
    riga = ws[ws.max_row]
    for i, cella in enumerate(riga):
        if i in _COLONNE_IMPORTO:
            cella.number_format = _FORMATO_IMPORTO
        elif i == _COLONNA_DATA and isinstance(cella.value, date):
            cella.number_format = _FORMATO_DATA


def genera_xlsx(conn, dal: str, al: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"
    ws.append(COLONNE)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in registro(conn, dal, al):
        _scrivi_riga(ws, r)
    # L'intestazione resta visibile scorrendo: un registro annuale e' lungo.
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Riepilogo IVA")
    ws2.append(["Aliquota %", "Imponibile", "ENPAV", "Base IVA", "IVA"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for g in riepilogo_iva(conn, dal, al):
        ws2.append([g["aliquota"], float(g["imponibile"]), float(g["enpav"]),
                    float(g["base_iva"]), float(g["iva"])])
        for cella in list(ws2[ws2.max_row])[1:]:
            cella.number_format = _FORMATO_IMPORTO
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def genera_zip_pdf(conn, dal: str, al: str, studio: dict) -> bytes:
    """ZIP con le copie PDF di tutti i documenti del periodo, note di credito incluse.

    Le note di credito servono al commercialista quanto le fatture: senza, la
    contabilita' del periodo non torna con il registro qui accanto.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for f in _documenti(conn, dal, al):
            fattura = leggi_fattura(conn, f["id"])
            gruppi = gruppi_iva_da_righe(fattura["righe"], fattura["enpav_pct"])
            pdf = genera_pdf_fattura(fattura, studio, gruppi)
            nome = f"{f['tipo_documento']}_{f['numero_visualizzato'].replace('/', '-')}.pdf"
            z.writestr(nome, pdf)
    return buf.getvalue()
